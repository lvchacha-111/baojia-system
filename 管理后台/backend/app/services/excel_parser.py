import pandas as pd
from openpyxl import load_workbook
import json
from pathlib import Path


class ExcelParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)

    def parse_to_json(self) -> dict:
        """解析 Excel 文件并转换为 JSON 格式"""
        data = {
            "version": "2025.11.1",
            "carriers": {
                "DHL": self._parse_dhl(),
                "UPS": self._parse_ups(),
                "FEDEX": self._parse_fedex()
            }
        }
        return data

    def _parse_dhl(self) -> dict:
        """解析 DHL 价格表"""
        sheet = self.wb["DHL"]

        price_table = {}
        for row_idx in range(4, 100):
            weight = sheet.cell(row_idx, 1).value
            if weight is None:
                break

            prices = {}
            for col_idx in range(3, 12):
                zone_num = col_idx - 2
                price = sheet.cell(row_idx, col_idx).value
                if price is not None:
                    prices[f"zone{zone_num}"] = float(price)

            price_table[str(float(weight))] = prices

        country_zones = {}
        for row_idx in range(3, 100):
            country = sheet.cell(row_idx, 13).value
            if country is None or country == "国家或地区":
                continue

            zone = sheet.cell(row_idx, 14).value
            peak_surcharge = sheet.cell(row_idx, 16).value

            if country and zone:
                country_zones[country] = {
                    "zone": int(zone),
                    "peak_surcharge": float(peak_surcharge) if peak_surcharge else 0
                }

        return {
            "price_table": price_table,
            "country_zones": country_zones,
            "size_limits": {
                "length": 300,
                "width": 200,
                "height": 180,
                "weight": 300
            },
            "parameters": {
                "fuel_rate": 1.4,
                "profit_rate": 1.2,
                "oversize_fee": 350,
                "overweight_fee": 750,
                "heavy_weight_threshold": 21,
                "special_countries": ["美国", "加拿大", "澳大利亚", "新西兰"]
            }
        }

    def _parse_ups(self) -> dict:
        return {
            "price_table": {},
            "country_zones": {},
            "size_limits": {"length": 300, "width": 200, "height": 180, "weight": 300},
            "parameters": {
                "fuel_rate": 1.4,
                "profit_rate": 1.2,
                "oversize_fee": 350,
                "overweight_fee": 750,
                "heavy_weight_threshold": 21
            }
        }

    def _parse_fedex(self) -> dict:
        return {
            "price_table": {},
            "country_zones": {},
            "size_limits": {"length": 300, "width": 200, "height": 180, "weight": 300},
            "parameters": {
                "fuel_rate": 1.4,
                "profit_rate": 1.2,
                "oversize_fee": 350,
                "overweight_fee": 750,
                "heavy_weight_threshold": 20.5
            }
        }


def parse_excel_to_json(excel_path: str, output_path: str):
    parser = ExcelParser(excel_path)
    data = parser.parse_to_json()

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return data
